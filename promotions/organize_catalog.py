# -*- coding: utf-8 -*-
"""
סידור קטלוג מבצעים לפי קטגוריות ראשיות — עם SUN/STAR ראשונים + סידור ידני.

לוקח גיליון קטלוג (כמו 'מבצעים ספטמבר 26'), מקבץ מחדש את הפריטים לפי
'קטגוריות ראשית', וכך שבתוך כל קטגוריה הפריטים המסומנים ב-priority = sun/star
מופיעים ראשונים. שאר הפריטים אחריהם.

סידור ידני (בנוסף לאוטומטי):
  • גיליון "סדר קטגוריות" — טבלה של קטגוריה + מספר סדר. הסקריפט קורא אותה
    וקובע לפיה את סדר הקטגוריות. ריק/חסר = סדר ההופעה המקורי.
  • עמודה "סדר פריט" בגיליון הראשי — מספר "נועץ" פריט במקום מסוים בתוך
    הקטגוריה שלו (לפני הסידור האוטומטי). ריק = אוטומטי.

בכל הרצה הפלט כולל את עמודת "סדר פריט" ואת גיליון "סדר קטגוריות" עם המספור
הנוכחי — כך אפשר לערוך אותם ולהריץ שוב, והסידור הידני יישמר.

דוגמה:
    python -m promotions.organize_catalog \
        --input  "קובץ המבצעים.xlsx" \
        --sheet  "מבצעים ספטמבר 26" \
        --out    "data/מבצעים ספטמבר מסודר.xlsx" \
        --month  ספטמבר
"""
from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E1F2")   # פס קטגוריה בהיר
FILL_PRIORITY = PatternFill("solid", fgColor="FFF2CC")   # רקע עדין ל-sun/star
FILL_MANUAL = PatternFill("solid", fgColor="FCE4D6")     # עמודת סידור ידני

PRIORITY_FIRST = {"sun", "star"}
CAT_HEADER = "קטגוריות ראשית"
PRI_HEADER = "priority"
KEY_HEADER = "מפתח פריט"
MANUAL_ITEM = "סדר פריט"      # עמודת סידור ידני של פריטים
CAT_ORDER_SHEET = "סדר קטגוריות"  # גיליון סידור ידני של קטגוריות
CAT_ORDER_NUM = "סדר"


def norm(v) -> str:
    return "" if v is None else str(v).replace("‏", "").strip()


def to_num(v):
    s = norm(v)
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_header_row(ws, key=KEY_HEADER, max_scan=15):
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if key in row_vals:
            return r
    raise SystemExit(f"לא נמצאה שורת כותרת עם '{key}' ב-{max_scan} השורות הראשונות")


def keep_columns(ws, header_row):
    """אינדקסי עמודות אמיתיות: כותרת לא-ריקה, לא כפילות '.1', ולא עמודת הסידור הידני."""
    keep = []
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(header_row, c).value)
        if h and not h.endswith(".1") and h != MANUAL_ITEM:
            keep.append(c)
    return keep


def read_manual_categories(wb):
    """קורא את גיליון 'סדר קטגוריות' -> {שם קטגוריה: מספר}. ריק אם אין."""
    if CAT_ORDER_SHEET not in wb.sheetnames:
        return {}
    ws = wb[CAT_ORDER_SHEET]
    # שורת כותרת = הראשונה שמכילה את שם עמודת הקטגוריה
    hrow = None
    for r in range(1, min(6, ws.max_row) + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if CAT_HEADER in vals or "קטגוריה" in vals:
            hrow = r
            break
    if hrow is None:
        return {}
    hdr = [norm(ws.cell(hrow, c).value) for c in range(1, ws.max_column + 1)]
    ci = next((i for i, h in enumerate(hdr) if h in (CAT_HEADER, "קטגוריה")), None)
    oi = next((i for i, h in enumerate(hdr) if h == CAT_ORDER_NUM), None)
    if ci is None or oi is None:
        return {}
    out = {}
    for r in range(hrow + 1, ws.max_row + 1):
        cat = norm(ws.cell(r, ci + 1).value)
        num = to_num(ws.cell(r, oi + 1).value)
        if cat and num is not None:
            out[cat] = num
    return out


def organize(input_path, sheet, add_visuals=True):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"גיליון '{sheet}' לא נמצא. גיליונות: {wb.sheetnames}")
    ws = wb[sheet]

    header_row = find_header_row(ws)
    cols = keep_columns(ws, header_row)
    headers = [norm(ws.cell(header_row, c).value) for c in cols]
    if CAT_HEADER not in headers:
        raise SystemExit(f"לא נמצאה עמודת '{CAT_HEADER}'. עמודות: {headers}")
    cat_pos = headers.index(CAT_HEADER)
    pri_pos = headers.index(PRI_HEADER) if PRI_HEADER in headers else None

    # עמודת סידור ידני של פריטים (אם קיימת בקלט)
    manual_item_col = None
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(header_row, c).value) == MANUAL_ITEM:
            manual_item_col = c
            break

    manual_cats = read_manual_categories(wb)  # {קטגוריה: מספר}

    # כותרות-על (מעל שורת הכותרת)
    top_rows = [[ws.cell(r, c).value for c in cols] for r in range(1, header_row)]

    # קריאת פריטים
    data, manual_items = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in cols]
        if all(v in (None, "") for v in vals):
            continue
        data.append(vals)
        manual_items.append(to_num(ws.cell(r, manual_item_col).value) if manual_item_col else None)

    # סדר קטגוריות: ידני (אם ניתן) קודם, אחרת סדר הופעה מקורי; #N/A/ריק בסוף
    cat_order, seen = [], set()
    for vals in data:
        c = norm(vals[cat_pos])
        if c not in seen:
            seen.add(c)
            cat_order.append(c)

    def cat_rank(cat):
        c = norm(cat)
        if c in ("", "#N/A", "#n/a", "None"):
            return (3, 0, 0)
        if c in manual_cats:                       # ידני מנצח
            return (0, manual_cats[c], 0)
        return (1, cat_order.index(c), 0)          # אחרת סדר מקורי

    INF = float("inf")

    def pri_rank(vals):
        p = norm(vals[pri_pos]).lower() if pri_pos is not None else ""
        return 0 if p in PRIORITY_FIRST else 1

    # מיון יציב: קטגוריה -> נעיצה ידנית של פריט -> sun/star -> סדר מקורי
    indexed = list(enumerate(zip(data, manual_items)))
    indexed.sort(key=lambda t: (
        cat_rank(t[1][0][cat_pos]),
        t[1][1] if t[1][1] is not None else INF,   # סדר פריט ידני
        pri_rank(t[1][0]),
        t[0],
    ))
    ordered = [(vals, man) for _, (vals, man) in indexed]

    # סדר קטגוריות סופי לפלט (לגיליון "סדר קטגוריות")
    final_cats = []
    for vals, _ in ordered:
        c = norm(vals[cat_pos])
        if c not in final_cats:
            final_cats.append(c)

    return {
        "headers": headers, "top_rows": top_rows, "ordered": ordered,
        "cat_pos": cat_pos, "pri_pos": pri_pos, "final_cats": final_cats,
        "manual_cats": manual_cats, "add_visuals": add_visuals,
    }


def write_xlsx(out_path, sheet_title, R):
    headers, top_rows, ordered = R["headers"], R["top_rows"], R["ordered"]
    cat_pos, pri_pos = R["cat_pos"], R["pri_pos"]
    add_visuals = R["add_visuals"]

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- גיליון ראשי ---
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.sheet_view.rightToLeft = True

    out_headers = [MANUAL_ITEM] + headers   # עמודת סידור ידני ראשונה
    r = 1
    for tr in top_rows:
        for c, v in enumerate([None] + tr, 1):
            ws.cell(r, c, v)
        ws.cell(r, 2).font = Font(bold=True)
        r += 1

    header_row_idx = r
    for c, h in enumerate(out_headers, 1):
        cell = ws.cell(r, c, h)
        cell.fill = FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(header_row_idx + 1, 1).coordinate
    r += 1

    prev_cat = object()
    for vals, man in ordered:
        cat = norm(vals[cat_pos])
        ws.cell(r, 1, man if man is not None else None)   # סדר פריט
        for c, v in enumerate(vals, 2):
            ws.cell(r, c, v)
        if add_visuals:
            ws.cell(r, 1).fill = FILL_MANUAL
            if cat != prev_cat:
                for c in range(1, len(out_headers) + 1):
                    ws.cell(r, c).fill = FILL_CATEGORY
                    ws.cell(r, c).font = Font(bold=True)
                prev_cat = cat
            elif pri_pos is not None and norm(vals[pri_pos]).lower() in PRIORITY_FIRST:
                ws.cell(r, pri_pos + 2).fill = FILL_PRIORITY
                ws.cell(r, pri_pos + 2).font = Font(bold=True)
        r += 1

    for c, h in enumerate(out_headers, 1):
        width = max([len(str(h))] + [len(str(ws.cell(rr, c).value or "")) for rr in range(header_row_idx, r)])
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 9), 45)

    # --- גיליון סידור קטגוריות ---
    cs = wb.create_sheet(CAT_ORDER_SHEET)
    cs.sheet_view.rightToLeft = True
    cs.cell(1, 1, CAT_HEADER).fill = FILL_HEADER
    cs.cell(1, 2, CAT_ORDER_NUM).fill = FILL_HEADER
    for c in (1, 2):
        cs.cell(1, c).font = Font(bold=True, color="FFFFFF")
        cs.cell(1, c).alignment = Alignment(horizontal="center")
    for i, cat in enumerate(R["final_cats"], start=1):
        cs.cell(i + 1, 1, cat)
        # שומר מספר ידני קיים אם היה, אחרת המיקום הנוכחי
        cs.cell(i + 1, 2, R["manual_cats"].get(cat, i))
    cs.column_dimensions["A"].width = 48
    cs.column_dimensions["B"].width = 10
    cs.freeze_panes = "A2"

    wb.save(out_path)
    return out_path, out_headers


def main(argv=None):
    p = argparse.ArgumentParser(description="סידור קטלוג מבצעים לפי קטגוריות עם sun/star ראשונים + סידור ידני")
    p.add_argument("--input", required=True, help="קובץ הקלט (XLSX)")
    p.add_argument("--sheet", required=True, help="שם הגיליון לסידור")
    p.add_argument("--out", required=True, help="נתיב פלט (XLSX חדש)")
    p.add_argument("--month", default="", help="שם החודש לכותרת הגיליון")
    p.add_argument("--no-visuals", action="store_true", help="בלי צביעה/פסי קטגוריה — רק סידור")
    args = p.parse_args(argv)

    R = organize(args.input, args.sheet, add_visuals=not args.no_visuals)
    title = f"מבצעים {args.month} מסודר" if args.month else f"{args.sheet} מסודר"
    out, out_headers = write_xlsx(args.out, title, R)

    rows = R["ordered"]
    cats = Counter(norm(v[R["cat_pos"]]) for v, _ in rows)
    pri_pos = R["pri_pos"]
    sunstar = sum(1 for v, _ in rows if pri_pos is not None and norm(v[pri_pos]).lower() in PRIORITY_FIRST)
    pinned = sum(1 for _, m in rows if m is not None)
    print(f"✅ נכתב: {out}")
    print(f"   {len(rows)} פריטים · {len(cats)} קטגוריות · {sunstar} מסומנים sun/star")
    print(f"   סידור ידני: {len(R['manual_cats'])} קטגוריות ממוספרות · {pinned} פריטים נעוצים")
    print(f"   גיליון '{CAT_ORDER_SHEET}' + עמודת '{MANUAL_ITEM}' מוכנים לעריכה")


if __name__ == "__main__":
    main()
