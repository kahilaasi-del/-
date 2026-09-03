# -*- coding: utf-8 -*-
"""
בניית קובץ מבצעים חודשי מסודר — פקודה אחת.

הרעיון: כל חודש נותנים רשימת מבצעים פשוטה (שם / מפתח פריט / אחוז מבצע),
והסקריפט מעשיר אותה אוטומטית מקטלוג המוצרים לפי מפתח פריט (ברקוד, מחיר,
קטגוריה, שם רשמי), משווה לחודש הקודם, וכותב אקסל מסודר עם צביעה:

    🟩 ירוק  = נשאר   (היה בחודש הקודם וגם החודש)
    🟥 אדום  = ירד    (היה בחודש הקודם, לא החודש)
    🟨 צהוב  = חדש    (חדש החודש, לא היה קודם)

דוגמה:
    python -m promotions.build_promos \
        --current   promotions/examples/מבצעים_אוקטובר.csv \
        --catalog   ~/downloads/קטלוג_מוצרים.xlsx \
        --previous  promotions/examples/מבצעים_ספטמבר.csv \
        --out       "data/מבצעים אוקטובר מסודר.xlsx" \
        --month     אוקטובר

הקלט החודשי (--current / --previous) הוא קובץ פשוט עם 3 עמודות לפחות:
    שם מוצר | מפתח פריט | מבצע
הקטלוג (--catalog) הוא ייצוא מחשבשבת/H-BI. הסקריפט מזהה אוטומטית עמודות
נפוצות; אם השמות שונים אצלך — אפשר לעקוף עם הדגלים --catalog-key וכו'.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# צבעים לסטטוס
# ---------------------------------------------------------------------------
FILL_STAYED = PatternFill("solid", fgColor="C6EFCE")  # ירוק — נשאר
FILL_DROPPED = PatternFill("solid", fgColor="FFC7CE")  # אדום — ירד
FILL_NEW = PatternFill("solid", fgColor="FFEB9C")      # צהוב — חדש
FILL_HEADER = PatternFill("solid", fgColor="305496")   # כחול כותרת

STATUS_STAYED = "נשאר"
STATUS_DROPPED = "ירד"
STATUS_NEW = "חדש"

STATUS_FILL = {
    STATUS_STAYED: FILL_STAYED,
    STATUS_DROPPED: FILL_DROPPED,
    STATUS_NEW: FILL_NEW,
}

# ---------------------------------------------------------------------------
# זיהוי עמודות גמיש — שמות נפוצים בעברית/אנגלית
# ---------------------------------------------------------------------------
KEY_CANDIDATES = ["מפתח פריט", "מפתח", "קוד פריט", "קוד מוצר", "מק\"ט", "מקט", "item_key", "sku", "code"]
NAME_CANDIDATES = ["שם פריט", "שם מוצר", "תיאור פריט", "תיאור", "שם", "name", "description"]
PROMO_CANDIDATES = ["מבצע", "אחוז מבצע", "הנחה", "אחוז הנחה", "promo", "discount"]
BARCODE_CANDIDATES = ["ברקוד", "בר קוד", "barcode", "ean"]
PRICE_CANDIDATES = ["מחיר", "מחירון", "מחיר מחירון", "price"]
CATEGORY_CANDIDATES = ["קטגוריה", "קבוצה", "משפחה", "category", "group"]
IMAGE_CANDIDATES = ["תמונה", "קישור תמונה", "image", "image_url", "img"]


def _norm(s: str) -> str:
    return str(s).replace("‏", "").replace("‎", "").strip().lower()


def find_col(df: pd.DataFrame, candidates: list[str], override: str | None = None) -> str | None:
    """מאתר עמודה לפי רשימת שמות אפשריים (התאמה מדויקת ואז חלקית)."""
    if override:
        if override in df.columns:
            return override
        raise SystemExit(f"עמודה שצוינה ידנית לא נמצאה: '{override}'. עמודות קיימות: {list(df.columns)}")
    cols = {_norm(c): c for c in df.columns}
    for cand in candidates:  # התאמה מדויקת
        if _norm(cand) in cols:
            return cols[_norm(cand)]
    for cand in candidates:  # התאמה חלקית
        for norm_c, orig in cols.items():
            if _norm(cand) in norm_c:
                return orig
    return None


def read_any(path: str | Path) -> pd.DataFrame:
    """קורא CSV או Excel לפי סיומת, עם טיפול בקידוד עברית."""
    path = Path(path)
    if not path.exists():
        raise SystemExit(f"קובץ לא נמצא: {path}")
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, dtype=str)
    for enc in ("utf-8-sig", "utf-8", "cp1255", "iso-8859-8"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise SystemExit(f"לא הצלחתי לקרוא את הקובץ (קידוד לא מוכר): {path}")


def clean_key(series: pd.Series) -> pd.Series:
    """מנרמל מפתח פריט למחרוזת נקייה להשוואה (ללא רווחים / .0 מספרי)."""
    s = series.astype(str).str.replace("‏", "", regex=False).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)  # 115600.0 -> 115600
    return s


# ---------------------------------------------------------------------------
# הלוגיקה המרכזית
# ---------------------------------------------------------------------------
def build(current_path, catalog_path, previous_path, month, overrides) -> pd.DataFrame:
    cur = read_any(current_path)
    cur_key = find_col(cur, KEY_CANDIDATES, overrides.get("current_key"))
    cur_name = find_col(cur, NAME_CANDIDATES)
    cur_promo = find_col(cur, PROMO_CANDIDATES)
    if not cur_key:
        raise SystemExit(f"לא נמצאה עמודת מפתח פריט בקובץ המבצעים. עמודות: {list(cur.columns)}")
    cur["_key"] = clean_key(cur[cur_key])

    # קטלוג המוצרים — מקור ההעשרה
    cat = read_any(catalog_path)
    cat_key = find_col(cat, KEY_CANDIDATES, overrides.get("catalog_key"))
    if not cat_key:
        raise SystemExit(f"לא נמצאה עמודת מפתח פריט בקטלוג. עמודות: {list(cat.columns)}")
    cat["_key"] = clean_key(cat[cat_key])
    cat = cat.drop_duplicates("_key", keep="first")

    enrich_map = {
        "שם מוצר": find_col(cat, NAME_CANDIDATES),
        "ברקוד": find_col(cat, BARCODE_CANDIDATES),
        "מחיר": find_col(cat, PRICE_CANDIDATES),
        "קטגוריה": find_col(cat, CATEGORY_CANDIDATES),
        "תמונה": find_col(cat, IMAGE_CANDIDATES),
    }

    # חודש קודם — לחישוב סטטוס
    prev_keys: set[str] = set()
    if previous_path:
        prev = read_any(previous_path)
        prev_key = find_col(prev, KEY_CANDIDATES, overrides.get("previous_key"))
        if not prev_key:
            raise SystemExit(f"לא נמצאה עמודת מפתח פריט בקובץ החודש הקודם. עמודות: {list(prev.columns)}")
        prev["_key"] = clean_key(prev[prev_key])
        prev_keys = set(prev["_key"])

    cur_keys = set(cur["_key"])

    # ---- בניית השורות ----
    rows = []

    # 1) פריטי החודש (נשאר / חדש)
    for _, r in cur.iterrows():
        k = r["_key"]
        status = STATUS_STAYED if k in prev_keys else STATUS_NEW
        crow = cat[cat["_key"] == k]
        row = {"מפתח פריט": k}
        row["שם מוצר"] = (r[cur_name] if cur_name else "") or ""
        for out_col, src_col in enrich_map.items():
            val = ""
            if src_col and not crow.empty:
                val = crow.iloc[0][src_col]
            if out_col == "שם מוצר" and not str(row["שם מוצר"]).strip():
                row["שם מוצר"] = val or ""
            elif out_col != "שם מוצר":
                row[out_col] = val if pd.notna(val) else ""
        row["מבצע " + month] = (r[cur_promo] if cur_promo else "") or ""
        row["בקטלוג"] = "לא נמצא" if crow.empty else ""
        row["סטטוס"] = status
        rows.append(row)

    # 2) פריטים שירדו (היו קודם, לא החודש)
    if previous_path:
        for _, r in prev[~prev["_key"].isin(cur_keys)].iterrows():
            k = r["_key"]
            prev_name = find_col(prev, NAME_CANDIDATES)
            prev_promo = find_col(prev, PROMO_CANDIDATES)
            crow = cat[cat["_key"] == k]
            row = {"מפתח פריט": k}
            row["שם מוצר"] = (r[prev_name] if prev_name else "") or ""
            for out_col, src_col in enrich_map.items():
                if out_col == "שם מוצר":
                    if not str(row["שם מוצר"]).strip() and src_col and not crow.empty:
                        row["שם מוצר"] = crow.iloc[0][src_col] or ""
                    continue
                val = crow.iloc[0][src_col] if (src_col and not crow.empty) else ""
                row[out_col] = val if pd.notna(val) else ""
            row["מבצע " + month] = ""  # ירד — אין מבצע החודש
            row["מבצע קודם"] = (r[prev_promo] if prev_promo else "") or ""
            row["בקטלוג"] = "לא נמצא" if crow.empty else ""
            row["סטטוס"] = STATUS_DROPPED
            rows.append(row)

    out = pd.DataFrame(rows)

    # סדר עמודות קבוע ונעים לעין
    preferred = ["סטטוס", "שם מוצר", "מפתח פריט", "ברקוד", "מחיר", "קטגוריה",
                 "מבצע " + month, "מבצע קודם", "תמונה", "בקטלוג"]
    ordered = [c for c in preferred if c in out.columns]
    ordered += [c for c in out.columns if c not in ordered]
    out = out[ordered]

    # מיון: נשאר, חדש, ירד — ובתוך כל קבוצה לפי קטגוריה/שם
    order = {STATUS_STAYED: 0, STATUS_NEW: 1, STATUS_DROPPED: 2}
    out["_o"] = out["סטטוס"].map(order).fillna(9)
    sort_cols = ["_o"] + [c for c in ("קטגוריה", "שם מוצר") if c in out.columns]
    out = out.sort_values(sort_cols).drop(columns="_o").reset_index(drop=True)
    return out


def write_xlsx(df: pd.DataFrame, out_path: str | Path, month: str) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        df.to_excel(xl, index=False, sheet_name=f"מבצעים {month}")
        ws = xl.sheets[f"מבצעים {month}"]
        ws.sheet_view.rightToLeft = True  # גיליון מימין לשמאל לעברית

        status_idx = list(df.columns).index("סטטוס") + 1

        # כותרת
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

        # צביעת שורות לפי סטטוס
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=status_idx).value
            fill = STATUS_FILL.get(status)
            if fill:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill

        # רוחב עמודות אוטומטי (מוגבל)
        for i, col in enumerate(df.columns, start=1):
            width = max(len(str(col)), *(len(str(v)) for v in df[col].astype(str))) if len(df) else len(str(col))
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 42)

    # סיכום מודפס
    counts = df["סטטוס"].value_counts().to_dict()
    print(f"✅ נכתב: {out_path}")
    print(f"   🟩 נשאר: {counts.get(STATUS_STAYED, 0)}  |  "
          f"🟨 חדש: {counts.get(STATUS_NEW, 0)}  |  "
          f"🟥 ירד: {counts.get(STATUS_DROPPED, 0)}")
    if "בקטלוג" in df.columns:
        missing = (df["בקטלוג"] == "לא נמצא").sum()
        if missing:
            print(f"   ⚠️  {missing} פריטים לא נמצאו בקטלוג (יש שם+מפתח, השאר להשלמה ידנית)")


def main(argv=None):
    p = argparse.ArgumentParser(description="בניית קובץ מבצעים חודשי מסודר")
    p.add_argument("--current", required=True, help="רשימת המבצעים של החודש (CSV/XLSX)")
    p.add_argument("--catalog", required=True, help="קטלוג המוצרים לייצוא/העשרה (CSV/XLSX)")
    p.add_argument("--previous", help="רשימת המבצעים של החודש הקודם (לחישוב נשאר/ירד/חדש)")
    p.add_argument("--out", required=True, help="נתיב פלט ה-XLSX")
    p.add_argument("--month", default="החודש", help="שם החודש לכותרת (למשל: אוקטובר)")
    p.add_argument("--catalog-key", help="שם עמודת מפתח פריט בקטלוג (עקיפה ידנית)")
    p.add_argument("--current-key", help="שם עמודת מפתח פריט בקובץ המבצעים (עקיפה ידנית)")
    p.add_argument("--previous-key", help="שם עמודת מפתח פריט בחודש הקודם (עקיפה ידנית)")
    args = p.parse_args(argv)

    overrides = {
        "catalog_key": args.catalog_key,
        "current_key": args.current_key,
        "previous_key": args.previous_key,
    }
    df = build(args.current, args.catalog, args.previous, args.month, overrides)
    write_xlsx(df, args.out, args.month)


if __name__ == "__main__":
    main()
