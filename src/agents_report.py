"""ניתוח כל הסוכנים מקובץ מכירות אחד (H-BI) — סטטוס מחושב מהמכירות.

מיועד לקובץ מכירות שמכיל את *כל* הסוכנים (עמודות: חתך/שם סוכן/שם חשבון/
מפתח חשבון/חודשים/כמות/מכירות/רווח). מפצל לפי חתך (סוכן), מחשב לכל לקוח
סטטוס פעילות מהמכירות, ומפיק חוברת Excel: 'השוואת סוכנים' + גליון לכל סוכן.

הערה: קובץ זה אינו כולל עיר/כתובת/אזור — לשכבה הגיאוגרפית יש להשתמש ב-hbi_report
עם קובץ 'מסלול מועשר' לכל סוכן.

הרצה:
    python -m src.agents_report --sales כל_הסוכנים.csv --out data/agents
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd

# חלונות מגמה (ניתן לכוונן): "אחרון" מול "קודם"
RECENT_MONTHS = [6, 7, 8]
EARLIER_MONTHS = [3, 4, 5]
ACTIVE = {"בעלייה", "יציב", "בירידה"}
RISK = {"בסכנת נטישה", "נטש"}
MONTH_HE = {1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל", 5: "מאי", 6: "יוני",
            7: "יולי", 8: "אוגוסט", 9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר"}
# יום בשבוע (pandas: שני=0..ראשון=6) → עברית
DOW_HE = {6: "יום א'", 0: "יום ב'", 1: "יום ג'", 2: "יום ד'", 3: "יום ה'", 4: "יום ו'", 5: "שבת"}
DAY_ORDER = ["יום א'", "יום ב'", "יום ג'", "יום ד'", "יום ה'", "יום ו'", "שבת", "לא משויך"]


def _compute_status(last_m: int, recent: float, earlier: float, current_month: int) -> str:
    """סטטוס פעילות מחושב מהמכירות בלבד."""
    if last_m <= current_month - 5:          # לא קנה ~5 חודשים
        return "נטש"
    if last_m <= current_month - 2:          # לא קנה 2-4 חודשים
        return "בסכנת נטישה"
    if earlier == 0 and recent > 0:
        return "בעלייה"
    if earlier > 0:
        t = (recent - earlier) / earlier
        if t > 0.10:
            return "בעלייה"
        if t < -0.10:
            return "בירידה"
    return "יציב"


def build_customer_table(sales_csv: str, current_month: int | None = None) -> pd.DataFrame:
    """מסכם לרמת סוכן×לקוח עם סטטוס מחושב."""
    df = pd.read_csv(sales_csv, dtype=str)
    for col, new in [("מכירות", "sales"), ("רווח", "profit"), ("כמות", "qty")]:
        if col in df.columns:
            df[new] = pd.to_numeric(df[col].str.replace(",", "", regex=False), errors="coerce").fillna(0)
    df["m"] = df["חודשים"].str[:2].astype(int)
    df["key"] = df["מפתח חשבון"].astype(str).str.strip()
    df["agent"] = df["שם סוכן"].astype(str).str.strip()
    df["seg"] = df["חתך"].astype(str).str.strip()
    cur = current_month or int(df["m"].max())

    recs = []
    for (seg, agent, key, name), grp in df.groupby(["seg", "agent", "key", "שם חשבון"]):
        recs.append({
            "seg": seg, "agent": agent, "key": key, "name": name,
            "sales": grp["sales"].sum(), "profit": grp["profit"].sum(),
            "last_m": int(grp["m"].max()), "months": grp["m"].nunique(),
            "recent": grp[grp["m"].isin(RECENT_MONTHS)]["sales"].sum(),
            "earlier": grp[grp["m"].isin(EARLIER_MONTHS)]["sales"].sum(),
        })
    c = pd.DataFrame(recs)
    c["margin"] = np.where(c["sales"] > 0, (c["profit"] / c["sales"] * 100).round(1), 0)
    c["last_month_name"] = c["last_m"].map(MONTH_HE)
    c["months_since"] = cur - c["last_m"]   # חודשים מאז הקנייה האחרונה
    c["status"] = c.apply(lambda r: _compute_status(r["last_m"], r["recent"], r["earlier"], cur), axis=1)
    c["active"] = c["status"].isin(ACTIVE)
    c["risk"] = c["status"].isin(RISK)
    return c


def parse_visits(xlsx_path: str, today: pd.Timestamp | None = None) -> pd.DataFrame:
    """מסכם דוח ביקורים (אסקי) לרמת לקוח: עיר, ביקור אחרון, מס' ביקורים, תדירות.

    מצפה לעמודות: לקוח (עם מספר בתחילת המחרוזת), עיר, תאריך.
    """
    today = today or pd.Timestamp.today().normalize()
    v = pd.read_excel(xlsx_path, header=0)
    v.columns = [str(c).strip() for c in v.columns]
    v["key"] = v["לקוח"].astype(str).str.extract(r"^\s*(\d+)")
    v["date"] = pd.to_datetime(v["תאריך"], errors="coerce", dayfirst=True)
    v = v.dropna(subset=["key", "date"])
    v["dow"] = v["date"].dt.dayofweek

    def _gap(dates: pd.Series) -> float:
        d = dates.sort_values()
        return round((d.max() - d.min()).days / (len(d) - 1), 1) if len(d) > 1 else np.nan

    def _city(x: pd.Series) -> str:
        val = x.mode().iloc[0] if len(x.mode()) else x.iloc[0]
        return re.sub(r"\s+", " ", re.sub(r"[0-9\-]", "", str(val))).strip() or "—"

    g = v.groupby("key").agg(city=("עיר", _city), last_visit=("date", "max"),
                             n_visits=("date", "size"),
                             dow=("dow", lambda x: int(x.mode().iloc[0]))).reset_index()
    g = g.merge(v.groupby("key")["date"].apply(_gap).rename("visit_gap").reset_index(), on="key", how="left")
    g["days_since_visit"] = (today - g["last_visit"]).dt.days
    g["visit_month"] = g["last_visit"].dt.month
    g["visit_day"] = g["dow"].map(DOW_HE)
    g["last_visit"] = g["last_visit"].dt.strftime("%d/%m/%Y")
    return g.drop(columns=["dow"])


def enrich_with_visits(c: pd.DataFrame, visits: pd.DataFrame) -> pd.DataFrame:
    """מוסיף לטבלת הלקוחות את נתוני הביקורים לפי מפתח לקוח + דגלי טיפול."""
    m = c.merge(visits, on="key", how="left")
    # דגל "ביקר ולא הזמין": ביקור אחרון מאוחר מחודש הקנייה האחרון
    m["visited_no_order"] = np.where(
        m["visit_month"].notna() & (m["visit_month"] > m["last_m"]), "כן", "")
    m["city"] = m["city"].fillna("—")
    m["last_visit"] = m["last_visit"].fillna("—")
    m["visit_day"] = m["visit_day"].fillna("לא משויך")
    m["n_visits"] = m["n_visits"].fillna(0).astype(int)
    m["days_since_visit"] = m["days_since_visit"].apply(lambda x: int(x) if pd.notna(x) else "")
    m["visit_gap"] = m["visit_gap"].apply(lambda x: x if pd.notna(x) else "")
    return m.drop(columns=["visit_month"])


def write_excel(c: pd.DataFrame, path: Path, min_customers: int = 10) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    HDR = PatternFill("solid", fgColor="2A78D6")
    HF = Font(color="FFFFFF", bold=True, size=11)
    BORD = Border(bottom=Side(style="thin", color="DDDDDD"))
    HDR2 = PatternFill("solid", fgColor="C55A11")  # כותרת הגליון הדחוף
    FLAG = PatternFill("solid", fgColor="FFF2CC")   # דגל "ביקר ולא הזמין"
    STCOL = {"בעלייה": "C6EFCE", "יציב": "C6EFCE", "בירידה": "FFEB9C",
             "בסכנת נטישה": "FCD5B4", "נטש": "FFC7CE"}
    W = {"name": 34, "key": 9, "sales": 13, "profit": 12, "margin": 8, "last_m": 9,
         "last_month_name": 13, "months_since": 15, "months": 11, "status": 13,
         "agent": 12, "seg": 7, "customers": 9, "active": 8, "risk": 8, "visit_day": 10,
         "city": 14, "last_visit": 12, "days_since_visit": 14, "n_visits": 10,
         "visit_gap": 11, "visited_no_order": 13}
    has_visits = "city" in c.columns

    def emit(ws, df, cols, headers, money=(), color_status=False, hdr_fill=HDR):
        ws.sheet_view.rightToLeft = True
        for j, h in enumerate(headers, 1):
            cc = ws.cell(1, j, h)
            cc.fill, cc.font = hdr_fill, HF
            cc.alignment = Alignment(horizontal="center", wrap_text=True)
        for i, (_, r) in enumerate(df.iterrows(), 2):
            for j, col in enumerate(cols, 1):
                v = r[col]
                if col in ("sales", "profit", "customers", "active", "risk"):
                    v = int(v)
                if col == "visit_day":
                    v = str(v)
                cc = ws.cell(i, j, v)
                cc.border = BORD
                if col in money:
                    cc.number_format = "#,##0"
                if color_status and col == "status" and r["status"] in STCOL:
                    cc.fill = PatternFill("solid", fgColor=STCOL[r["status"]])
                if col == "visited_no_order" and v == "כן":
                    cc.fill = FLAG
        for j, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = W.get(col, 12)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    ag = c.groupby(["seg", "agent"]).agg(customers=("key", "count"), active=("active", "sum"),
                                         risk=("risk", "sum"), sales=("sales", "sum"),
                                         profit=("profit", "sum")).reset_index().sort_values("sales", ascending=False)
    ag["active"] = ag["active"].astype(int)
    ag["risk"] = ag["risk"].astype(int)
    ag["margin"] = np.where(ag["sales"] > 0, (ag["profit"] / ag["sales"] * 100).round(1), 0)
    if has_visits:
        ag["cities"] = c.groupby(["seg", "agent"])["city"].nunique().reindex(
            list(zip(ag["seg"], ag["agent"]))).values
        emit(wb.create_sheet("השוואת סוכנים"), ag,
             ["seg", "agent", "customers", "cities", "active", "risk", "sales", "profit", "margin"],
             ["חתך", "סוכן", "לקוחות", "ערים", "פעילים", "בסיכון", "מכירות", "רווח", "% רווח"],
             money=("sales", "profit"))
        # 🚩 גליון דחוף: ביקר ולא הזמין (הסוכן ביקר, הלקוח לא הזמין מאז)
        vno = c[c["visited_no_order"] == "כן"].sort_values("sales", ascending=False)
        emit(wb.create_sheet("🚩 ביקר ולא הזמין"), vno,
             ["agent", "name", "key", "city", "sales", "last_month_name", "months_since",
              "last_visit", "days_since_visit", "status"],
             ["סוכן", "שם לקוח", "מס לקוח", "עיר", "מכירות", "חודש קנייה אחרון",
              "חודשים ללא קנייה", "ביקור אחרון", "ימים מאז ביקור", "סטטוס"],
             money=("sales",), color_status=True, hdr_fill=HDR2)
        # לקוחות להעברה (נטשו / בסכנת נטישה)
        trans = c[c["risk"]].sort_values(["agent", "sales"], ascending=[True, False])
        emit(wb.create_sheet("לקוחות להעברה"), trans,
             ["agent", "name", "key", "city", "sales", "last_month_name", "months_since",
              "visited_no_order", "last_visit", "status"],
             ["סוכן", "שם לקוח", "מס לקוח", "עיר", "מכירות", "חודש קנייה אחרון",
              "חודשים ללא קנייה", "ביקר ולא הזמין", "ביקור אחרון", "סטטוס"],
             money=("sales",), color_status=True)
        c = c.copy()
        c["_dayord"] = c["visit_day"].map({d: i for i, d in enumerate(DAY_ORDER)}).fillna(9)
        cols = ["name", "key", "city", "visit_day", "sales", "profit", "margin", "last_month_name",
                "months_since", "visited_no_order", "last_visit", "days_since_visit",
                "n_visits", "visit_gap", "status"]
        heads = ["שם לקוח", "מס לקוח", "עיר", "יום ביקור", "מכירות", "רווח", "% רווח",
                 "חודש קנייה אחרון", "חודשים ללא קנייה", "ביקר ולא הזמין", "ביקור אחרון",
                 "ימים מאז ביקור", "מס' ביקורים", "תדירות", "סטטוס"]
        sort_by = ["_dayord", "city", "sales"]
        sort_asc = [True, True, False]
    else:
        emit(wb.create_sheet("השוואת סוכנים"), ag,
             ["seg", "agent", "customers", "active", "risk", "sales", "profit", "margin"],
             ["חתך", "סוכן", "לקוחות", "פעילים", "בסיכון", "מכירות", "רווח", "% רווח"],
             money=("sales", "profit"))
        cols = ["name", "key", "sales", "profit", "margin", "last_month_name", "months_since", "months", "status"]
        heads = ["שם לקוח", "מס לקוח", "מכירות", "רווח", "% רווח", "חודש אחרון", "חודשים מאז קנייה", "חודשים פעיל", "סטטוס"]
        sort_by, sort_asc = ["sales"], [False]

    used = set()
    for _, a in ag.iterrows():
        sub = c[(c["seg"] == a["seg"]) & (c["agent"] == a["agent"])].sort_values(sort_by, ascending=sort_asc)
        if len(sub) < min_customers:
            continue
        code = str(a["seg"]).replace("^", "") or "?"
        nm = f"{code}-{a['agent']}"[:31]
        base, k = nm, 1
        while nm in used:
            k += 1
            nm = f"{base[:28]}_{k}"
        used.add(nm)
        emit(wb.create_sheet(nm), sub, cols, heads, money=("sales", "profit"), color_status=True)
    wb.save(path)


def main() -> None:
    p = argparse.ArgumentParser(description="ניתוח כל הסוכנים מקובץ מכירות")
    p.add_argument("--sales", required=True, help="קובץ מכירות CSV (כל הסוכנים)")
    p.add_argument("--visits", default=None, help="דוח ביקורים XLSX (אסקי) — להעשרה בעיר/ביקורים")
    p.add_argument("--out", default="data/agents", help="תיקיית פלט")
    p.add_argument("--current-month", type=int, default=None)
    p.add_argument("--min-customers", type=int, default=10, help="מינימום לקוחות לגליון סוכן נפרד")
    args = p.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    c = build_customer_table(args.sales, args.current_month)
    if args.visits:
        c = enrich_with_visits(c, parse_visits(args.visits))
        print(f"  הועשר מדוח ביקורים: {int((c['city'] != '—').sum())} לקוחות עם עיר")
    c.to_csv(out / "customers.csv", index=False, encoding="utf-8-sig")
    write_excel(c, out / "all_agents.xlsx", args.min_customers)
    n_agents = c.groupby(["seg", "agent"]).ngroups
    print(f"✓ {n_agents} סוכנים · {len(c)} רשומות לקוח")
    print(f"  פעילים: {int(c['active'].sum())} | בסיכון: {int(c['risk'].sum())}")
    print(f"  פלטים ב-{out}/ : customers.csv · all_agents.xlsx")


if __name__ == "__main__":
    main()
