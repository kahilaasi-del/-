"""דוח סוכן מקבצי H-BI של חשבשבת — צינור מלא לשימוש חוזר.

מקבל שני קבצים שמייצאים מ-H-BI:
  1. קובץ מכירות (CSV) — מכירות חודשיות לפי לקוח (עם עמודת "מפתח חשבון").
  2. קובץ מסלול מועשר (XLSX) — לקוחות עם עיר/כתובת/אזור/סטטוס (עם "מס' לקוח").

מייצר תחת תיקיית הפלט:
  - merged.csv          : נתונים מאוחדים ונקיים
  - report.xlsx         : חוברת עבודה לשטח (5 גליונות)
  - dashboard.html      : דשבורד ויזואלי עצמאי (נפתח בדפדפן)

הרצה:
    python -m src.hbi_report --sales מכירות.csv --route מסלול.xlsx --out data/agent31
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd

# סטטוסים לקטגוריות
ACTIVE_STATUS = {"בעלייה", "יציבים", "חדשים", "בירידה"}
RISK_STATUS = {"בסכנת נטישה", "נטשו לאחרונה", "אבודים", "נטשו", "לא הזמינו אף פעם"}
DAY_ORDER = ["יום א'", "יום ב'", "יום ג'", "יום ד'", "יום ה'", "יום ו'"]

# מיפוי עמודות בגליון "מסלול מועשר" (לפי מיקום)
ROUTE_COLS = {
    "order": 0, "name": 1, "city": 2, "addr": 3, "region": 4, "status": 5,
    "trend": 6, "lastvisit": 7, "freq": 8, "days": 9, "rec": 10, "mode": 11, "key": 12,
}


def parse_route(xlsx_path: str, sheet: str = "מסלול מועשר") -> pd.DataFrame:
    """מפענח את גליון המסלול: כותרות יום ('יום א') ואזור ('» עיר') הן שורות-חתך."""
    raw = pd.read_excel(xlsx_path, sheet_name=sheet, header=None)
    rows, day, grp = [], None, None
    for _, r in raw.iterrows():
        c0 = str(r[0]).strip()
        if c0.startswith("יום"):
            day = c0.split("(")[0].strip()
            continue
        if c0.startswith("»"):
            grp = c0.replace("»", "").strip()
            continue
        if c0 in ("סדר", "nan", "", "None"):
            continue
        if re.fullmatch(r"\d+(\.0)?", c0):  # שורת לקוח אמיתית
            key = str(r[ROUTE_COLS["key"]]).replace(".0", "").strip()
            if key in ("nan", ""):
                continue
            rows.append({
                "day": day, "group": grp, "key": key,
                "name": str(r[ROUTE_COLS["name"]]).strip(),
                "city": str(r[ROUTE_COLS["city"]]).strip(),
                "addr": str(r[ROUTE_COLS["addr"]]).strip(),
                "region": str(r[ROUTE_COLS["region"]]).strip(),
                "status": str(r[ROUTE_COLS["status"]]).strip(),
                "lastvisit": str(r[ROUTE_COLS["lastvisit"]]).strip(),
                "freq": str(r[ROUTE_COLS["freq"]]).strip(),
            })
    return pd.DataFrame(rows)


def parse_sales(csv_path: str) -> pd.DataFrame:
    """מסכם מכירות חודשיות לרמת לקוח: סה\"כ מכירות/רווח/כמות, חודש אחרון, מס' חודשים פעילים."""
    s = pd.read_csv(csv_path, dtype=str)
    for col, new in [("מכירות", "sales"), ("רווח", "profit"), ("כמות", "qty")]:
        if col in s.columns:
            s[new] = pd.to_numeric(s[col].str.replace(",", "", regex=False), errors="coerce").fillna(0)
    s["m"] = s["חודשים"].str[:2].astype(int)
    s["key"] = s["מפתח חשבון"].astype(str).str.strip()
    return s.groupby("key").agg(
        sales=("sales", "sum"), profit=("profit", "sum"), qty=("qty", "sum"),
        last_m=("m", "max"), months=("m", "nunique"),
    ).reset_index()


def merge(route: pd.DataFrame, sales: pd.DataFrame, current_month: int) -> pd.DataFrame:
    """מאחד מסלול+מכירות לפי מפתח לקוח ומחשב דגלי פעילות."""
    m = route.merge(sales, on="key", how="left")
    for col in ["sales", "profit", "qty", "last_m", "months"]:
        m[col] = m[col].fillna(0)
    m["last_m"] = m["last_m"].astype(int)
    m["months"] = m["months"].astype(int)
    m["margin"] = np.where(m["sales"] > 0, (m["profit"] / m["sales"] * 100).round(1), 0)
    # פעיל לפי מכירות = קנה ב-60 יום האחרונים (חודש נוכחי או קודם)
    m["active_recency"] = m["last_m"] >= current_month - 1
    m["active_status"] = m["status"].isin(ACTIVE_STATUS)
    m["risk"] = m["status"].isin(RISK_STATUS)
    return m


# ─────────────────────── Excel ───────────────────────
def write_excel(m: pd.DataFrame, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    HDR = PatternFill("solid", fgColor="2A78D6")
    HF = Font(color="FFFFFF", bold=True, size=11)
    BORD = Border(bottom=Side(style="thin", color="DDDDDD"))
    STCOL = {"בעלייה": "C6EFCE", "יציבים": "C6EFCE", "חדשים": "C6EFCE", "בירידה": "FFEB9C",
             "בסכנת נטישה": "FCD5B4", "נטשו לאחרונה": "FFC7CE", "אבודים": "FFC7CE",
             "נטשו": "FFC7CE", "לא הזמינו אף פעם": "F2F2F2"}
    W = {"name": 34, "city": 13, "addr": 30, "region": 17, "status": 13, "sales": 12,
         "profit": 11, "margin": 8, "day": 9, "group": 18, "freq": 8, "last_m": 9,
         "lastvisit": 13, "key": 9}

    def sheet(name, df, cols, headers, money=()):
        ws = wb.create_sheet(name)
        ws.sheet_view.rightToLeft = True
        for j, h in enumerate(headers, 1):
            c = ws.cell(1, j, h)
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center", vertical="center")
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, col in enumerate(cols, 1):
                v = row[col]
                if col in ("sales", "profit", "qty"):
                    v = int(v)
                c = ws.cell(i, j, v)
                c.border = BORD
                if col in money:
                    c.number_format = "#,##0"
                if col == "status" and row["status"] in STCOL:
                    c.fill = PatternFill("solid", fgColor=STCOL[row["status"]])
        for j, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = W.get(col, 12)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    # סקירה
    ov = wb.create_sheet("סקירה")
    ov.sheet_view.rightToLeft = True
    ov.column_dimensions["A"].width = 28
    ov.column_dimensions["B"].width = 16
    ov["A1"] = "סקירת סוכן"
    ov["A1"].font = Font(bold=True, size=14)
    kpis = [("לקוחות במסלול", len(m)), ("פעילים — סטטוס", int(m["active_status"].sum())),
            ("פעילים — 60 יום", int(m["active_recency"].sum())), ("בסיכון / נטישה", int(m["risk"].sum())),
            ("לא הזמינו אף פעם", int((m["sales"] == 0).sum())), ("סה\"כ מכירות ₪", int(m["sales"].sum())),
            ("סה\"כ רווח ₪", int(m["profit"].sum())), ("אזורים", m["region"].nunique()),
            ("ערים", m["city"].nunique())]
    for i, (k, v) in enumerate(kpis, 3):
        ov.cell(i, 1, k).font = Font(bold=True)
        c = ov.cell(i, 2, v)
        if "₪" in k:
            c.number_format = "#,##0"

    r = m.groupby("region").agg(customers=("key", "count"), active=("active_recency", "sum"),
                                risk=("risk", "sum"), sales=("sales", "sum"),
                                profit=("profit", "sum")).reset_index().sort_values("sales", ascending=False)
    r["active"] = r["active"].astype(int)
    r["risk"] = r["risk"].astype(int)
    sheet("לפי אזור", r, ["region", "customers", "active", "risk", "sales", "profit"],
          ["אזור", "לקוחות", "פעילים", "בסיכון", "מכירות", "רווח"], money=("sales", "profit"))

    risk = m[m["risk"]].sort_values("sales", ascending=False)
    sheet("בסיכון לפעולה", risk, ["name", "key", "status", "region", "city", "addr", "sales", "last_m", "lastvisit"],
          ["שם לקוח", "מס לקוח", "סטטוס", "אזור", "עיר", "כתובת", "מכירות 2026", "חודש אחרון", "ביקור אחרון"], money=("sales",))

    route = m.sort_values(["day", "group", "sales"], ascending=[True, True, False])
    sheet("מסלול לפי יום", route, ["day", "group", "name", "key", "city", "addr", "status", "sales", "freq"],
          ["יום", "אזור מסלול", "שם לקוח", "מס לקוח", "עיר", "כתובת", "סטטוס", "מכירות", "תדירות"], money=("sales",))

    allc = m.sort_values("sales", ascending=False)
    sheet("כל הלקוחות", allc, ["name", "key", "region", "city", "addr", "status", "sales", "profit", "margin", "last_m"],
          ["שם לקוח", "מס לקוח", "אזור", "עיר", "כתובת", "סטטוס", "מכירות", "רווח", "% רווח", "חודש אחרון"],
          money=("sales", "profit"))
    wb.save(path)


# ─────────────────────── Dashboard JSON ───────────────────────
def build_dashboard_data(m: pd.DataFrame) -> dict:
    out: dict = {}
    out["kpi"] = {
        "customers": int(len(m)), "active_status": int(m["active_status"].sum()),
        "active_recency": int(m["active_recency"].sum()), "risk": int(m["risk"].sum()),
        "never": int((m["sales"] == 0).sum()), "sales": int(m["sales"].sum()),
        "profit": int(m["profit"].sum()), "regions": int(m["region"].nunique()),
        "cities": int(m["city"].nunique()),
    }
    r = m.groupby("region").agg(customers=("key", "count"), active=("active_recency", "sum"),
                                risk=("risk", "sum"), sales=("sales", "sum")).reset_index().sort_values("sales", ascending=False)
    out["regions"] = [{"region": x.region, "customers": int(x.customers), "active": int(x.active),
                       "risk": int(x.risk), "sales": int(x.sales)} for x in r.itertuples()]
    st = m["status"].value_counts()
    out["status"] = [{"status": k, "count": int(v)} for k, v in st.items()]
    rl = m[m["risk"]].sort_values("sales", ascending=False)
    out["risk_list"] = [{"name": x.name, "city": x.city, "addr": x.addr, "region": x.region,
                         "status": x.status, "sales": int(x.sales)} for x in rl.itertuples()]
    ci = m.groupby("city").agg(customers=("key", "count"), sales=("sales", "sum")).reset_index().sort_values("sales", ascending=False)
    out["cities"] = [{"city": x.city, "customers": int(x.customers), "sales": int(x.sales)} for x in ci.itertuples()]
    dd = m.groupby("day").agg(customers=("key", "count"), active=("active_recency", "sum"),
                              risk=("risk", "sum"), sales=("sales", "sum")).reset_index()
    dd["o"] = dd["day"].apply(lambda x: DAY_ORDER.index(x) if x in DAY_ORDER else 9)
    dd = dd.sort_values("o")
    out["days"] = [{"day": x.day, "customers": int(x.customers), "active": int(x.active),
                    "risk": int(x.risk), "sales": int(x.sales)} for x in dd.itertuples()]
    return out


def write_dashboard(data: dict, path: Path, title: str) -> None:
    """כותב דשבורד HTML עצמאי (התבנית ב-dashboard/hbi_template.html)."""
    template = Path(__file__).resolve().parent.parent / "dashboard" / "hbi_template.html"
    html = template.read_text(encoding="utf-8")
    html = html.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    html = html.replace("__TITLE__", title)
    path.write_text(html, encoding="utf-8")


def main() -> None:
    p = argparse.ArgumentParser(description="דוח סוכן מ-H-BI")
    p.add_argument("--sales", required=True, help="קובץ מכירות CSV")
    p.add_argument("--route", required=True, help="קובץ מסלול מועשר XLSX")
    p.add_argument("--out", default="data/agent_report", help="תיקיית פלט")
    p.add_argument("--title", default="דוח סוכן — לקוחות לפי אזור וכתובת")
    p.add_argument("--current-month", type=int, default=None,
                   help="חודש נוכחי (1-12) להגדרת 'פעיל'; ברירת מחדל: החודש המקסימלי בנתונים")
    args = p.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    route = parse_route(args.route)
    sales = parse_sales(args.sales)
    cur = args.current_month or int(sales["last_m"].max())
    m = merge(route, sales, cur)

    m.to_csv(out / "merged.csv", index=False, encoding="utf-8-sig")
    write_excel(m, out / "report.xlsx")
    data = build_dashboard_data(m)
    write_dashboard(data, out / "dashboard.html", args.title)

    print(f"✓ נותחו {len(m)} לקוחות (חודש נוכחי={cur:02d})")
    print(f"  פעילים(60 יום): {int(m['active_recency'].sum())} | בסיכון: {int(m['risk'].sum())}")
    print(f"  פלטים ב-{out}/ : merged.csv · report.xlsx · dashboard.html")


if __name__ == "__main__":
    main()
